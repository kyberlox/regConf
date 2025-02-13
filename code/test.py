import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

from sqlalchemy import create_engine
from sqlalchemy import select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy import create_engine, MetaData, Column, Integer, Text, Float

engine = create_engine('postgresql+psycopg2://kyberlox:4179@postgres/pdb')

class Base(DeclarativeBase): pass

'''
class Table2(Base):
    __tablename__ = 'table2'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)

class Table10(Base):
    __tablename__ = 'table10'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)
'''

class pakingParams(Base):
    __tablename__ = 'paking_params'
    id = Column(Integer, primary_key=True)
    mark = Column(Text, nullable=True)
    DN = Column(Float, nullable=True)
    PN = Column(Float, nullable=True)
    M = Column(Float, nullable=True)
    S = Column(Float, nullable=True)

Base.metadata.create_all(bind=engine)
SessionLocal = sessionmaker(autoflush=True, bind=engine)
db = SessionLocal()

def migration():
    wb3 = load_workbook("./paking_params.xlsx")
    sheet = wb3["result"]

    pak_res = {"added" : [], "exists" : []}
    for i in range(2, sheet.max_row):
        mark = str(sheet[f"A{i}"].value)
        DN = float(sheet[f"B{i}"].value)
        PN = float(sheet[f"C{i}"].value)
        M = sheet[f"D{i}"].value
        S = float(sheet[f"E{i}"].value)
        if type(M) == type("") or type(M) == type(1.0):
            M = float(sheet[f"D{i}"].value)
        else:
            M = None
        print(mark, DN, PN, M, S)

        example = pakingParams(mark=mark, DN=DN, PN=PN, M=M, S=S)
        
        request = db.query(pakingParams).filter(pakingParams.mark == mark, pakingParams.DN == DN, pakingParams.PN == PN).first()
        
        #если нет - добавить
        if request == None:
            db.add(example)
            db.commit()

            curr = {
                "№" : i,
                "ID" : example.id,  
                "mark" : mark, 
                "DN" : DN, 
                "PN" : PN,
                "M" : M,
                "S" : S
            }
            pak_res["added"].append(curr)
        #если есть - пропустить
        else:
            curr = {
                "№" : i,
                "ID" : example.id,  
                "mark" : mark, 
                "DN" : DN, 
                "PN" : PN,
                "M" : M,
                "S" : S
            }
            pak_res["exists"].append(curr)
    return pak_res

print(migration())