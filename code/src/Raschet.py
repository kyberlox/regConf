import math
from math import sqrt, log, exp, pi, log10

#В каких величинах PN? кгс/см2 =>  0.098067 * PN МПа
from sqlalchemy import create_engine
from sqlalchemy import select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy import create_engine, MetaData, Column, Integer, Text, Float

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

import os
from dotenv import load_dotenv



load_dotenv()

user = os.getenv('user')
pswd = os.getenv('pswd')
port = os.getenv('PORT')

engine = create_engine(f'postgresql+psycopg2://{user}:{pswd}@postgres/pdb')

class Base(DeclarativeBase): pass

class Params(Base):
    __tablename__ = 'parametrs_table'
    id = Column(Integer, primary_key=True)
    DNS = Column(Float, nullable=True)
    Pnd = Column(Text, nullable=True)
    DN = Column(Float, nullable=True)
    PN = Column(Float, nullable=True)
    spring_material = Column(Text, nullable=True)
    spring_number = Column(Text, nullable=True)
    valve_type = Column(Text, nullable=True)

class Table2(Base):
    __tablename__ = 'table2'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)

class Table10(Base):
    __tablename__ = 'table10'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)

class pakingParams(Base):
    __tablename__ = 'paking_params'
    id = Column(Integer, primary_key=True)
    mark = Column(Text, nullable=True)
    DN = Column(Float, nullable=True)
    PN = Column(Float, nullable=True)
    M = Column(Float, nullable=True)
    S = Column(Float, nullable=True)

Base.metadata.create_all(bind=engine)
SessionLocal = sessionmaker(autoflush=True, bind=engine)
db = SessionLocal()



def linear_interpolation(x1, y1, x2, y2, x):
    """
    Функция для линейной интерполяции.

    :param x1: x-координата первой точки
    :param y1: y-координата первой точки
    :param x2: x-координата второй точки
    :param y2: y-координата второй точки
    :param x: x-координата точки, для которой нужно найти y
    :return: интерполированное значение y
    """
    if x1 == x2:
        raise ValueError("x1 и x2 не должны быть равны, чтобы избежать деления на ноль.")

    # Вычисляем y по формуле линейной интерполяции
    y = y1 + (x - x1) * (y2 - y1) / (x2 - x1)
    return y



def searchT2(T, Pn):
    #print(f"T2: {Pn}")
    #найти все подходящие строки их DNS и P1 - больше искомых
    request = db.query(Table2).filter(Table2.T >= T, Table2.Pn >= Pn).all()

    if request == None or len(request) == 0:
        return False
    ans = False

    #найти самый подходящий - MIN по DNS и P1
    minT = request[0].T
    minPn = request[0].Pn
    for example in request:
        if (example.T <= minT) and (example.Pn <= minPn):
            minT = example.T
            minPn = example.Pn
            ans = {
                "ID" : example.id,  
                "T" : example.T, 
                "Pn" : example.Pn * 10, 
                "PN" : example.P
            }
    #print(ans)
    return ans
    
def searchT10(T, Pn):
    #print(f"T10: {Pn}")
    #найти все подходящие строки их DNS и P1 - больше искомых
    request = db.query(Table10).filter(Table10.T >= T, Table10.Pn >= Pn).all()

    if request == None or len(request) == 0:
        return False
    ans = False

    #найти самы подходящий - MIN по DNS и P1
    minT = request[0].T
    minPn = request[0].Pn
    for example in request:
        if (example.T <= minT) and (example.Pn <= minPn):
            minT = example.T
            minPn = example.Pn
            ans = {
                "ID" : example.id,  
                "T" : example.T, 
                "Pn" : example.Pn, 
                "PN" : example.P
            }
    #print(ans)
    return ans

def searchParams(DNS, Pn, PN, valve_type):
    print(DNS, PN, valve_type)
    #найти все подходящие строки их DNS и P1 - больше искомых
    request = db.query(Params).filter(Params.DNS >= DNS, Params.PN == PN, Params.valve_type == valve_type).all()

    if request == None or request == []:
        return False
    ans = False

    #найти самый подходящий - MIN по DNS и P1
    minDNS = request[0].DNS
    #minP1 = request[0].P1
    minPN = request[0].PN
    ##print("###")
    for example in request:

        #print(example.id, example.DNS, example.valve_type, example.DN, example.PN)

        try:
            Pn1 = str(example.Pnd).split("...")[0]
            Pn2 = str(example.Pnd).split("...")[1]
            #print(Pn1, Pn2)

            #print(f"example.DNS <= minDNS {example.DNS <= minDNS} example.PN == minPN {example.PN == minPN} float(Pn1) <= Pn <= float(Pn2) {float(Pn1)} {Pn} {float(Pn2)} {float(Pn1) <= Pn <= float(Pn2)}")
            if (example.DNS <= minDNS)  and (example.PN == minPN) and (float(Pn1) <= Pn <= float(Pn2)):
                minDNS = example.DNS
                #minP1 = example.P1
                minPN = example.PN
                ans = {
                    "ID" : example.id,
                    "DNS" : example.DNS,
                    "Pnd" : example.Pnd,
                    "DN" : example.DN,
                    "PN" : example.PN,
                    "spring_material" : example.spring_material,
                    "spring_number" : example.spring_number,
                    "valve_type" : valve_type
                }
            elif (Pn <= float(Pn2)) and (Pn <= 4) and (example.DNS <= minDNS)  and (example.PN == minPN):
                minDNS = example.DNS
                # minP1 = example.P1
                minPN = example.PN
                ans = {
                    "ID": example.id,
                    "DNS": example.DNS,
                    "Pnd": example.Pnd,
                    "DN": example.DN,
                    "PN": example.PN,
                    "spring_material": example.spring_material,
                    "spring_number": example.spring_number,
                    "valve_type": valve_type
                }
        except:
            print("###")
            print(example.id)
            print(example.Pnd)
            print("###")
    #print("###")
    #print(ans)

    return ans

def get_by_mark(mark, DN, PN):
    mark = mark[2:5]
    request = db.query(pakingParams).filter(pakingParams.mark == mark, pakingParams.DN == DN, pakingParams.PN == PN).first()
    print(mark, DN, PN)
    if request is None:
        return ("Нет данных", "Нет данных")
    else:

        M = "Нет данных"
        S = "Нет данных"
        if request.M is not None:
            M = request.M
        elif request.S is not None:
            S = request.S
        print(M, S)
        return (M, S)



data_mean = {
    "environment_id": "ID среды",
    "name": "Название среды",
    "environment": "Состояние (Жидкость/Газ)",
    "molecular_weight": "Молекулярная масса",
    "density": "Плотность",
    "material": "Материал",
    "viscosity": "Вязкость",
    "isobaric_capacity": "Удельная изобарная теплоемиокость",
    "molar_mass": "Молярная масса",
    "isochoric_capacity": "Удельная изохорная теплоемиокость",
    "adiabatic_index": "Показатель адиабаты",
    "compressibility_factor": "Фактор сжимаемости",
    "mark": "Маркировка",

    "valve_type": "Тип предохранительного клапана Пружинный или Пилотный (B/H)",
    "force_open": "Устройство принудительного открытия [Да/Нет]",
    "Pn": "Давление настройки (МПа)",
    "Pp":  "Проитводавление (МПа)",
    "Pp_din": "Динамическое противодавление (МПа)",
    "Gab": "Максимальный аварийный расход жидкости и газа (кг/час)",
    "N": "Количество параллельно установленных и одновременно работающих клапанов (шт)",
    "pre_Kc": "Мембранно-предохранительное устройство установлено до и/или после",
    "T": "Температура рабочей среды",
    "climate": "Климатическое исполнение по ГОСТ 15150-69 (У1 ХЛ1 УХЛ1 М1)",

    "T_min": "минимальная температура",
    "T_max": " максимальная температура",
    "Pno": "Давление начала открытия с противодавлением",
    "Ppo": "Давление полного открытия с противодавлением",
    "P1": "Давление на входе",
    "P2": "Давление на выходе",
    "Kw": "Коэффициент, учитывающий эффект неполного открытия разгруженных ПК из-за противодавления",
    "Gideal": "Массовая скорость",
    "pre_DN": "DN предварительный",
    "DN_s": "Диаметр седла клапана",
    "Pnd": "Диапазон давлений настройки",
    "DN": "Номиннальный диаметр",
    "PN": "Номинальное давление",
    "DN2": "Номиннальный иаметр на выходе",
    "PN2": "Номинальное давление на выходе",
    "spring_material": "Материал пружины",
    "spring_number": "Номер пружины",
    "color": "Покраска",

    "need_bellows": "Переменное противодавление или необходим сильфон на пружинные ПК по требованию ОЛ",
    "contact_type": "Тит котакта (металл-металл / металл-неметалл)",
    "tightness": "Герметичность затвора по ГОСТ 9544-2015 (A AA B C)",
    "open_close_type": "Открытый / Закрытый тип",
    "joining_standard": "Исполнение фланцев на входе по ГОСТ 33259-2015 или Исполнение фланцев на входе по ГОСТ 33259-2015",
    "joining_type": "Тип присоединения [Фланцевое, Под приварку, Штуцерно-торцовое, Муфтовое, Ниппельное, Кламповое, Комбинированное]",
    "inlet_flange": "Фланец на входе (C D E F J)",
    "outlet_flange": "Фланец на входе (C D E F J)",
    "detonation_node": "Узел подрыва недоступен для заказа",
    "material_bellows": "Материал сильфона",
    "material_spool": "Материал золотника",
    "material_saddle": "Материал седла",
    "weight": "Маccа",
    "painting_area": "Площадь под покраску",
    "packaging": "Стандартная (клапан упакован в стрейч-пленку, на паллет). Хранение под навесом.",  # Упаковка",
    "trials": "Испытания",

    "rotary_plugs": "Наличие поворотных заглушек",
    "thermal_cover": "Наличие термочехла",
    "abrasive_particles": "Наличие в рабочей среде абразивных частиц",
    "warranty": "Нарантийный срок службы, мес.",
    "assignment": "Назначенный срок службы, лет",
    "acceptance": "Приемка",
    "adapters": "Переходники",
    "needKOF": "Наличие КОФ [Да/Нет]",
    "need_ZIP": "Наличие ЗИП [Да/Нет]",
    "reciprocal_connections": "Комплект ответных присоединительных частей",
    "pipe_material": "Материал трубы Материал и размер трубопровода",
    "quantity": "Количество",
    "docs": " Документация",
    "additionally": "Дополнительно",
    "OL_num": "Номер отпросного листа",
}


def mixture(envs : list, climate : str, T : float):
    result = {
        "name" : "",
        "environment" : "",
        "molecular_weight" : 0,
        "density" : 0,
        "density_ns": 0,
        "material" : "",
        "viscosity" : 0,
        "isobaric_capacity" : 0,
        "molar_mass" : 0,
        "isochoric_capacity" : 0,
        "adiabatic_index" : 0,
        "compressibility_factor" : 1,
    }

    env_type = set()
    for env in envs:
        env_type.add(env["environment"])


    r_max = 0
    #проверка типа среды смеси
    if len(env_type) == 1: #если среда однородная
        result["environment"] = env_type.pop()
        if result["environment"] == "Жидкость": #если среда - жидкость
            ch_den = 0
            zn_den = 0
            pre_viscosity = 0
            for env in envs:
                r = env["r"]
                result["name"] += f"{env['name']}:{r*100}% "
                result["molecular_weight"] += env["molecular_weight"] * r
                ch_den += env["density"] * r
                zn_den += r
                pre_viscosity += log10(env["viscosity"]) * r

                '''
                if r > r_max:
                    # Плотность несущей среды
                    r_max = r
                    result["density_ns"] = env["density"]
                '''


            result["density"] = ch_den/zn_den
            result["density_ns"] = result["density"]
            result["viscosity"] = 10**(pre_viscosity)
            
            '''
            #добавить подбор материала result["material"]
            if "Морская вода" in result["name"]:
                result["material"] = "12Х18Н12М3ТЛ"
            elif ("Масло подсолнечное" in result["name"]) or ("Лимонная кислота" in result["name"]) or ("Молочная кислота" in result["name"]):
                result["material"] = "12Х18Н9ТЛ"
            '''

        elif result["environment"] == "Газ": #если среда - газ
            viscosity_сh = 0
            viscosity_zn = 0
            pre_M = 0
            adiabatic_index = 0
            adiabatic_index_zn = 0
            density_ns_zn = 0
            for env in envs:
                r = env["r"]
                result["name"] += f"{env['name']}:{r*100}% "
                M_i = env["molar_mass"]
                u_i = env["viscosity"]
                pre_M += M_i * r
                viscosity_сh += u_i * r * sqrt(M_i)
                viscosity_zn += r * sqrt(M_i)
                adiabatic_index += env['adiabatic_index'] * r

                # плотность при н.у.
                result["density_ns"] += (M_i * r)
                #density_ns_zn += r
                '''
                if r > r_max:
                    # Плотность несущей среды
                    r_max = r
                    result["density_ns"] = ((env["molar_mass"] / 22.4) * r) / r
                '''
                
            result["molar_mass"] = pre_M #/100
            result["viscosity"] = viscosity_сh / viscosity_zn
            result["adiabatic_index"] = adiabatic_index

            # плотность при н.у.
            result["density_ns"] = result["density_ns"] / 22.4
            result["density"] = None
            #result["density"] = result["density_ns"] * (273.15) / (T + 273.15)

            print()


    else:
        result["environment"] = "Смесь"
        density_ch = 0
        density_zn = 0
        pre_u = 0
        for env in envs:

            r = env["r"]
            result["name"] += f"{env['name']}:{r*100}% "

            # pre_viscosity += log10(env["viscosity"]) * r

            if env["environment"] == "Газ":
                M = env["molar_mass"]
                density_ch += (env["molar_mass"] / 22.4) * r
                density_zn += r
            elif env["environment"] == "Жидкость":
                M = env["molecular_weight"]
                density_ch += env["density"] * r
                density_zn += r

            pre_u += r * env["viscosity"] * M

            if r > r_max:
                # Плотность несущей среды при нормальных условиях
                r_max = r
                result["density_ns"] = density_ch / density_zn

        #рабочая плотность
        result["density"] = density_ch / density_zn
        result["viscosity"] = pre_u
        # result["viscosity"] = 10**(pre_viscosity)


    material = []
    for env in envs:
        if env['name'] == 'Сероводород' and r < 0.06 and result["environment"] == "Смесь":
            material.append(f"25Л")
        else:
            material.append(env['material'])

    ln = 0
    for mat in material:
        if len(mat) > ln:
            ln = len(mat)
            result["material"] = mat

    #если климатика => то материал
    if ((climate == "ХЛ1") or (climate == "УХЛ1")) and (result["material"] == "25Л"):
        result["material"] = "20ГЛ"

    result["T"] = T

    return result

'''def density_true(data):
    T = data["T"]
    Pn = data["Pn"]
    density_ns = data["density_ns"]
    P0 = 0.101325

    if Pn <= 0.3:
        Ppo = Pn + 0.05
    elif (Pn > 0.3) and (Pn <= 6):
        Ppo = 1.15 * Pn
    elif Pn > 6:
        Ppo = 1.1 * Pn

    # плотность рабочая Менделлева_Клайперона
    result["density"] = density_ns * (Ppo * 273.15) / (0.101325 * (T + 273.15))

    return result'''

def Raschet(dt):
    kys = ["viscosity", "Pn", "Pp", "Pp_din", "Gab", "N", "pre_Kc", "density", "climate", "material", "environment"]
    for param in kys:
        if param not in dt:
            return {"err": f"Не корректно определён или не орпеделён параметр \'{data_mean[param]}\' "}

    P_atm = 0.101320
    R = 8.31446261815324  # Газовая постоянная ( Па / (моль * K))

    u = dt["viscosity"]
    Pn = dt["Pn"]
    Pp = dt["Pp"]
    Pp_din = dt["Pp_din"]
    Gab = dt["Gab"]
    N = dt["N"]
    pre_Kc = dt["pre_Kc"]



    climate = dt["climate"]
    model = {
        "У1": [-40, 40],
        "ХЛ1": [-60, 40],
        "УХЛ1": [-60, 40],
        "М1": [-40, 40]
    }

    T_min, T_max = model[climate]

    T = dt["T"]

    if pre_Kc:
        Kc = 0.9
    else:
        Kc = 1

    # n = #Введите показатель изоэнтропы среды

    # Давление начала открытия
    # Давление полного открытия
    if Pn <= 0.3:
        Pno = Pn + 0.02
        Ppo = Pn + 0.05
    elif (Pn > 0.3) and (Pn <= 6):
        Pno = 1.07 * Pn
        Ppo = 1.15 * Pn
    elif Pn > 6:
        Pno = 1.05 * Pn
        Ppo = 1.1 * Pn
    else:
        return {"err": f"Невозможно определить давление начала открытия и давление полного открытия, при давлении настройки = {Pn}"}

    if dt["environment"] == "Газ":
        #p1 = dt["density_ns"]  * (Ppo * 273.15) / (0.101325 * (T + 273.15))
        p1 = (dt["density_ns"] * Ppo * 100000 * 273.15) / (101325 * (T + 273.15) * 8.314)

        if  dt["convertGab"]:# и размерность м3/час
            # Домножить Gab
            Gab *= p1
    else:
        p1 = dt["density"]
    dt["density"] = p1

    # Максимально допустимое давление аварийного сброса;
    P_ab_max = 1.1 * Pno

    # Абсолютное давление до клапана,
    P1 = Ppo + P_atm  # < P_ab_max

    # Абсолютное давление за клапаном при его полном открытии
    P2 = Pp + P_atm  # = P_sbr

    # Отношение абсолютных давлений;
    B = P2 / P1

    if dt["environment"] == "Газ":

        if "molar_mass" not in dt:
            return {"err": f"Не корректно определён или не орпеделён параметр \'f{data_mean['molar_mass']}\' "}

        M = dt["molar_mass"]

        p1 = P1 * 1000 * M / (R * (T + 273.15))

        alpha = 0.8
        # (Д.22)
        if (Ppo / Pn) == 1.1:
            if (Pp / Pno) <= 0.3:
                Kw = 1
            else:
                Kw = 1.1027 + 0.4007 * (Pp / Pno) - 2.4577 * (Pp / Pno) ** 2
        # (Д.23)
        elif (Ppo / Pn) == 1.15:
            if (Pp / Pno) <= 0.37:
                Kw = 1
            else:
                Kw = 1.2857 - 0.7603 * (Pp / Pno)
        # (Д.24)
        elif ((Ppo / Pn) > 1.2) and ((Pp / Pno) >= 0.5):
            Kw = 1
        # (Д.25)
        elif ((Ppo / Pn) > 1.1) and ((Ppo / Pn) <= 1.15):
            # Kw определяют линейной интерполяцией по (Ppo / Pn) между значениями, полученными по (Д.22) и (Д.23)
            #return {"err": f"Нет возможности расчитать {data_mean['Kw']} для соотношения 1.1 < Ppo / Pn <= 1.15, при {data_mean['Ppo']} = {Ppo} и {data_mean['Pn']} = {Pn} "}

            # (Д.22)
            Ppo_Pn_1 = 1.1
            Kw_1 = 1.1027 + 0.4007 * (Pp / Pno) - 2.4577 * (Pp / Pno) ** 2

            # (Д.23)
            Ppo_Pn_2 = 1.15
            Kw_2 = 1.2857 - 0.7603 * (Pp / Pno)

            Kw = linear_interpolation(Ppo_Pn_1, Kw_1, Ppo_Pn_2, Kw_2, Ppo / Pn)

        # (Д.26)
        elif ((Ppo / Pn) > 1.15) and ((Ppo / Pn) <= 1.2):
            # Kw определяют линейной интерполяцией по (Ppo / Pn) между значениями, полученными по (Д.23) и (Д.24)
            #return {"err": f"Нет возможности расчитать {data_mean['Kw']} для соотношения 1.1 < Ppo / Pn <= 1.15, при {data_mean['Ppo']} = {Ppo} и {data_mean['Pn']} = {Pn} "}

            # (Д.23)
            Ppo_Pn_1 = 1.15
            Kw_1 = 1.2857 - 0.7603 * (Pp / Pno)

            # (Д.24)
            Ppo_Pn_2 = 1.21
            Kw_2 = 1

            Kw = linear_interpolation(Ppo_Pn_1, Kw_1, Ppo_Pn_2, Kw_2, Ppo / Pn)



        # показатель изоэнтропии
        # dt["isobaric_capacity"] / dt["isochoric_capacity"] = dt["adiabatic_index"]
        n = dt["adiabatic_index"]  # / dt["compressibility_factor"]

        Bkr = (2 / (n + 1)) ** (n / (n - 1))
        # определим режим истечения
        if B <= Bkr:  #
            print('критический режим истечения')
            Kb = 1
            if n == 1:
                Kp_kr = 0.60653 ** 2
            else:
                # Kp_kr = n*(Bkr**((n+1)/n)) #на самом деле, тут корень, но его будем извлекать в конце
                Kp_kr = sqrt((2 * n) / (n + 1)) * (2 / (n + 1)) ** (1 / (n - 1))  # или можно так
        else:  # докритический режим
            print('докритический режим истечения')
            Kp_kr = 1
            if n == 1:
                Kb = B ** 2 * -2 * exp * log(B)  # на самом деле, тут корень, но его будем извлекать в конце
            else:
                Kb = (((n + 1) / (n - 1)) * (B ** (2 / n) - B ** ((n + 1) / n)) * ((n + 1) / 2)) ** 2

        # P1 * p1
        Gideal = Kp_kr * Kb * sqrt(P1 * p1)

        if Gideal <= 0:
            return {
                "err": f"Одно из значений = 0:\n {data_mean['Ppo']} : {Ppo}\n {data_mean['Pno']} : {Pno}\n {Ppo}\n {data_mean['Kb']} : {Kb}\n {data_mean['Ppo']} : {Ppo}\n  {data_mean['Kp_kr']} : {Kp_kr}\n  {data_mean['Kw']} : {Kw}\n"}

    else:
        #(Д.21)
        alpha = 0.6
        # p1 = dt["density"]
        if (Pp / Pno) <= 1.15:
            Kw = 1
        elif ((Pp / Pno) > 1.15) and ((Pp / Pno) <= 0.25):
            Kw = 0.875 + 1.8333 * (Pp / Pno) - 6.6667 * (Pp / Pno) ** 2
        elif (Pp / Pno) > 0.25:
            Kw = 1.149 - 0.988 * (Pp / Pno)

        Kp = sqrt(2 * (1 - B))  # на самом деле, тут корень, но его будем извлекать в конце
        Gideal = Kp * sqrt(P1 * p1)

        if Gideal <= 0:
            return {"err": f"Одно из значений = 0:\n {data_mean['Ppo']} : {Ppo}\n {data_mean['Pno']} : {Pno}\n {Ppo}\n {data_mean['Pn']} : {Pn}\n {data_mean['Pp']} : {Pp}\n {data_mean['Kw']} : {Kw}\n"}

    # print(Kp_kr, P1, p1)
    # print(Kp, P1, p1)
    # print(Gideal)

    DN_s = None
    pre_DN = 0
    Kv = 1

    while DN_s != pre_DN:

        pre_F = Gab / (3.6 * alpha * Kv * Kw * Kc * Gideal * N)
        print(pre_F, ":", Gab, alpha, Kv, Kw, Kc, Gideal, N)
        if pre_F == 0:
            return {"err": f"Одно из значений = 0:\n Kv : {Kv}\n {data_mean['Kw']} : {Kw}\n {data_mean['Kc']} : {Kc}\n {data_mean['Gideal']} : {Gideal}\n"}
        pre_DN = sqrt((4 * pre_F) / pi)

        Re = (Gideal * p1 * pre_DN) / u  # Gideal
        if (Re >= 1000) and (Re <= 100000):
            Kv = (0.9935 + (2.8780 / Re ** 0.5) + (342.75 / Re ** 1.5)) ** (-1)
        elif (Re < 1000):
            Kv = 0.975 * sqrt(1 / 170 / (Re + 0.98))
        else:
            Kv = 1

        F = Gab / (3.6 * alpha * Kv * Kw * Kc * Gideal * N)
        DN_s = sqrt((4 * F) / pi)
    DN_s = math.ceil(DN_s * 10) / 10

    # перевести из МПа в кгс/см2
    new_dt = {
        "T_min": T_min,  # Минимальная рабочая температура
        "T_max": T_max,  # Максивальная рабочая температура
        "Pno": Pno * 10.197162,  # Давление начала открытия с противодавлением
        "Ppo": Ppo * 10.197162,  # Давление полного открытия с противодавлением
        "P1": P1 * 10.197162,  # Давление на входе
        "P2": P2 * 10.197162,  # Давление на выходе
        "Kw": Kw,  # Коэффициент, учитывающий эффект неполного открытия разгруженных ПК из-за противодавления
        "Gideal": Gideal,  # Массовая скорость
        #"pre_DN": pre_DN,
        "pre_DN": DN_s  # DN предварительный
    }

    # Номинальное давление
    new_dt["PN"] = f"Невозмажно подобрать при сочитании параметров: \nТемпература рабочей среды = {T} \n Давление настройки = {Pn}"
    if dt["material"] == "20ГЛ" or dt["material"] == "25Л":
        ex = searchT2(T, Pn * 10.197162)
    else:
        ex = searchT10(T, Pn * 10.197162)

    if ex:
        PN = ex["PN"]
    else:
        return {"err": f"Нет возможности подобрать {data_mean['PN']}, при {data_mean['T']} = {T} и {data_mean['Pn']} = {Pn}"}
    # print("PN по T и Pn:", PN)

    # Деаметр ПК
    new_dt["DN"] = f"Невозмажно подобрать при сочитании параметров: \nДаметр седла клапана = {DN_s} \n Давление на входе = {PN}"
    example = searchParams(DN_s, Pn * 10.197162, PN, dt["valve_type"])


    if example:
        new_dt["DN_s"] = example["DNS"] # Номинальный диаметр седла
        new_dt["DN"] = example["DN"]    # Номинальный диаметр
        new_dt["PN"] = example["PN"]    # Номиннальное давление
        # print("PN по DN:", example["PN"])
    else:
        return {
            "err": f"Нет возможности подобрать {data_mean['DN']}, при {data_mean['DN_s']} = {DN_s}, {data_mean['Pn']} = {Pn}, {data_mean['PN']} = {PN} и {data_mean['valve_type']} = \'{dt['valve_type']}\' "}

    DN2 = {
        25.0: 40.0,
        50.0: 80.0,
        80.0: 100.0,
        100.0: 150.0,
        150.0: 200.0,
        200.0: 300.0
    }
    new_dt["DN2"] = DN2[new_dt["DN"]]

    PN2 = {
        16.0: 6,
        40.0: 16.0,
        63.0: 40.0,
        100.0: 40.0,
        160.0: 40.0,
        250.0: 40.0
    }
    new_dt["PN2"] = PN2[new_dt["PN"]]

    # Площадь седла клапана
    DN_s = example["DNS"]
    S = (pi * DN_s**2 )/ 4
    new_dt["S"] = S
    # Эффективная площадь седла калапан
    new_dt["S_eff"] = S * alpha

    new_dt["spring_material"] = example["spring_material"]
    new_dt["spring_number"] = example["spring_number"]
    new_dt["Pnd"] = example["Pnd"]

    # подбор сильфона !!!!!!!!!!!!!!!!!!!!! сильфон только на пружине
    if (dt["valve_type"] == 'В') and (((example["spring_material"] == '51ХФА') and (T > 120)) or ((example["spring_material"] == '50ХФА') and (T > 250))):
        new_dt["need_bellows"] = True
    elif dt["valve_type"] == 'В':
            new_dt["need_bellows"] = [True, False]
    elif dt["valve_type"] == 'Н':
        new_dt["need_bellows"] = False
    else:
        new_dt["need_bellows"] = False


    # окрытый закрытый тип
    env_name = dt["name"]
    env_names = []
    for ev in env_name.split():
        env_names.append(ev[:ev.find(":")])

    # вода агрессиваня?
    cool_env = ["Вода", "Водяной пар", "Воздух", "Азот", "Вода"]

    evil_env = False
    cool = 0
    for en in env_names:
        # убрать из смеси неагрессивные среды
        if en in cool_env:
            # print(en)
            cool += 1

    if cool == len(env_names) and (dt["valve_type"] == 'В') and (((example["spring_material"] == '51ХФА') and (T > 120)) or ((example["spring_material"] == '50ХФА') and (T > 250))):
        evil_env = True

    # print(evil_env)

    open_close_type = "закрытого типа"
    if evil_env and T :
        open_close_type = "открытого типа"
        dt["need_bellows"] = False

    dt["open_close_type"] = open_close_type  # открытый или закрытый тип

    if "Сероводород" in dt["name"] and "Хлор" in dt["name"] and PN >= 0.003:
            #молибденовое исполнение
            dt["material"] = "12Х18Н12М3ТЛ"

    all_dt = dt | new_dt
    return all_dt

def mark_params(dt):
    kys = ["valve_type", "PN", "PN2", "DN", "T", "joining_type", "need_bellows", "material", "mark"]
    for param in kys:
        if param not in dt:
            return {"err" : f"Не корректно определён или не орпеделён параметр \'{data_mean[param]}\'"}
        
    valve_type = dt["valve_type"]
    PN = dt["PN"]
    PN2 = dt["PN2"]
    DN = dt["DN"]
    T = dt["T"]
    Ppo = dt["Ppo"]
    joining_type = dt["joining_type"]

    err = False

    #тип контакта
    if valve_type == "В": #у пружинного - строго металл-металл
        contact_type = "металл-металл"

    elif (valve_type == "Н") and (PN > 160):  
        contact_type = "металл-металл"

    elif (valve_type == "Н") and (PN <= 160): #если пилотный - по умлочанию металл-неметалл, но можно выбрать
        contact_type = ["металл-неметалл", "металл-металл"] #можно заменить

    else:
        err = {"error" : "Невозможно определить тип контакта", "value" : f"Некорректое значение типа ПК: {valve_type}"}

    #подбор фланцев
    if joining_type == "Фланцевое":
        inlet_flange = ['B']#B C D F J K
        if PN == 6.0:
            inlet_flange = ['B', 'C', 'D', 'E', 'F']
        if PN == 16.0:
            inlet_flange = ['B', 'C', 'D', 'E', 'F']
        if PN == 40.0:
            inlet_flange = ['F', 'C', 'D', 'E']
        if PN == 63.0 or PN == 100.0 or PN == 160.0:
            inlet_flange = ['J', 'K', 'F', 'C', 'D', 'E']
        if PN == 250.0:
            inlet_flange = ['K', 'D']
        
        outlet_flange = ['B']#B C D F J K
        if PN2 == 6.0:
            outlet_flange = ['B', 'C', 'D', 'E', 'F']
        if PN2 == 16.0 or PN2 == 16.4:
            outlet_flange = ['B', 'C', 'D', 'E', 'F']
        if PN2 == 40.0:
            outlet_flange = ['F', 'C', 'D', 'E']
        if PN2 == 63.0 or PN2 == 100.0 or PN2 == 160.0:
            outlet_flange = ['J', 'K', 'F', 'C', 'D', 'E']
        if PN2 == 250.0:
            outlet_flange = ['K', 'D']

    else:
        inlet_flange = None
        outlet_flange = inlet_flange

    

    material_bellows = "08Х18Н10Т" if dt["need_bellows"] else ""



    # Испытания
    trials = "По ТУ"
    if "Сероводород" in dt["name"]:
        need = False
        need_M = False

        s = dt["name"]
        inx = s.rfind("Сероводород")
        s_new = s[inx + 12:]
        r = float(s_new[:s_new.find("%")])

        if r > 6.0 and dt["environment"] == "Смесь":
            need = True

        r *= 0.01
        if Ppo * r >= 0.003:
            need = True

        if need and "Хлор" in dt["name"]:
            need_M = True

        if need:
            # нержавеющее исполнение
            dt["material"] = "12Х18Н9ТЛ"
            trials = "По СТ ЦКБА 052-2008\n\nИспытания материала корпуса:\n 1) Хим. Состав \n 2) На растяжение при +20 град. С \n 3) KCU при -60 град. С \n 4) Твердость \n 5) Стойкость к МКК \n 6) ВИК \n 7) РК \n 8) Капиллярный контроль \n\nИспытания материала золотника и седла: \n 1) Хим. Состав \n 2) На растяжение при +20 град. С \n 3) Контроль неметаллических включений \n 4) Контроль макроструктуры \n 5) Твердость \n 6) Стойкость к МКК \n 7) ВИК \n 8) РК \n 9) Капиллярный контроль"
        if need_M:
            # молибденовое исполнение
            dt["material"] = "12Х18Н12М3ТЛ"
            trials = "По СТ ЦКБА 052-2008\n\nИспытания материала корпуса:\n 1) Хим. Состав \n 2) На растяжение при +20 град. С \n 3) KCU при -60 град. С \n 4) Твердость \n 5) Стойкость к МКК \n 6) ВИК \n 7) РК \n 8) Капиллярный контроль \n\nИспытания материала золотника и седла: \n 1) Хим. Состав \n 2) На растяжение при +20 град. С \n 3) Контроль неметаллических включений \n 4) Контроль макроструктуры \n 5) Твердость \n 6) Стойкость к МКК \n 7) ВИК \n 8) РК \n 9) Капиллярный контроль"



    if dt["material"] == "25Л" and T <= 200:
        material_spool = "20Х13"
    elif dt["material"] == "25Л" and T > 200:
        material_spool = "12Х18Н10Т"
    elif dt["material"] == "20ГЛ" and T > 200:
        material_spool = "12Х18Н10Т"
    elif dt["material"] == "20ГЛ" and T <= 200:
        material_spool = "14Х17Н2"
    elif dt["material"] == "12Х18Н9ТЛ" and T > 200:
        material_spool = "12Х18Н10Т"
    elif dt["material"] == "12Х18Н9ТЛ" and T <= 200:
        material_spool = "12Х18Н10Т"
    else:
        material_spool = "10Х17Н13М3Т"
    
    if dt["material"] == "25Л":
        color = [
            f"Серый RAL7035 по технологической инструкции 38877941.25206.01013 АО \"НПО Регулятор\" ", #Заводская
            f"Серый RAL7035 cистема АКП С4 по № П2-05 ТИ-0002", #Роснефть
            f"Красный RAL3020 по СТО Газпром 9.1-018-2012", #Газпром
            "Другое"
            ]
    elif dt["material"] == "20ГЛ":
        color = [
            f"Синий RAL5017 по технологической инструкции 38877941.25206.01013 АО \"НПО Регулятор\" ", #Заводская
            f"Синий RAL5017 система АКП С4 по № П2-05 ТИ-0002", #Роснефть
            f"Красный RAL3020 по СТО Газпром 9.1-018-2012", #Газпром
            "Другое"
            ]
    else:
        color = [
            f"Голубой RAL5012 по технологической инструкции 38877941.25206.01013 АО \"НПО Регулятор\" ", #Заводская
            f"Голубой RAL5012 система АКП С4 по № П2-05 ТИ-0002", #Роснефть
            f"Красный RAL3020 по СТО Газпром 9.1-018-2012", #Газпром
            "Другое"
            ]

    weight, painting_area = get_by_mark(dt["mark"], DN, PN)

    packaging = [
        "Упаковка на европаллет (1200х800)",
        "Упаковка груза в ящики из OSB по ТУ “АО НПО Регулятор”",
        "Упаковка груза в ящики из OSB по ТУ “АО НПО Регулятор”  в северном исполнении",
        "Упаковка груза в ящики из OSB по ТУ “АО НПО Регулятор”  в морском исполнении",
        "Упаковка груза в дощатые ящики по ГОСТ 10198",
        "Упаковка груза в дощатые ящики по ГОСТ 10198 в северном исполнении",
        "Упаковка груза в дощатые ящики по ГОСТ 10198 в морском исполнении"
        ]
    
    if DN <= 50:
        packaging.insert(1, "Пенная защитная упаковка груза")

    assignment = "25 лет" if dt["need_bellows"] else "30 лет"



    new_dt = {
        "material_bellows": material_bellows,   # Материал сильфона
        "material_spool": material_spool,       # Материал золотника
        "material_saddle": material_spool,      # Материал седла
        "weight": weight,                       # Маccа
        "color": color,                         # Цвет
        "painting_area": painting_area,         # Площадь под покраску
        "packaging": packaging,                 # Упаковка
        "trials": trials,                       # Испытания
        "assignment": assignment,               # Срок службы
    }
            
    #заполнеие параметров и выгрузка
    if err:
        return err
    else:
        dt["contact_type"] = contact_type       #тип присоединения
        dt["inlet_flange"] = inlet_flange       #варианты фланца на входе
        dt["outlet_flange"] = outlet_flange     #варианты фланца на выходе
        dt = dt | new_dt
        return dt

def get_tightness(dt):
    kys = ["valve_type", "contact_type", "DN"]
    for param in kys:
        if param not in dt:
            return {"err" : f"Не корректно определён или не орпеделён параметр \'{data_mean[param]}\'"}
            
    #класс герметичнности
    #бывает ли АА?
    tightness = []
    if dt["valve_type"] == "Н":
        if dt["contact_type"] == "металл-металл":
            tightness = ["С"]
        elif dt["contact_type"] == "металл-неметалл":
            tightness = ["В", "А", "АА", "С"]
    elif dt["valve_type"] == "В":
        if dt["DN"] == 25.0:
            tightness = ["В", "С"]
        else:
            tightness = ["В", "А", "С"]
    else:
        return {"error" : "Невозможно определить класс герметичнности", "value" : f"Некорректое значение типа ПК: {dt['valve_type']}"}

    #варианты класса герметичности
    dt["tightness"] = tightness

    return dt


def make_XL(dt):

    WB = load_workbook("./src/ТКП.xlsx")
    sheet = WB['Лист1']



    data_keys = {
        "B": "OL_num",
        "C": "mark",
        "F": "quantity",
        "G": "pipe_material",
        "H": "name",
        "I": "T",
        "L": "climate",
        "M": "force_open",
        "N": "need_bellows",
        "O": "DN",
        "P": "PN",
        "Q": "DN2",
        "R": "PN2",
        "T": "Gab",
        "U": "material",
        "V": "material_bellows",
        "W": "material_spool",
        "X": "material_saddle",
        "Y": "spring_material",
        "Z": "joining_type",
        "AA": "contact_type",
        "AB": "weight",
        "AC": "painting_area",
        "AD": "color",
        "AE": "tightness",
        "AF": "spring_number",
        "AG": "Pnd",
        "AK": "Pp",
        "AO": "needKOF",
        "AP": "need_ZIP",
        "AQ": "adapters",
        "AR": "thermal_cover",
        "AS": "docs",
        "AT": "pre_Kc",
        "AU": "rotary_plugs",
        "AV": "packaging",
        "AW": "acceptance",
        "AX": "trials",
        "AZ": "assignment",
        "BA": "additionally"
    }

    for i, position in enumerate(dt, start=3):
        # проверка
        kys = list(data_keys.values())
        kys += ["valve_type", "open_close_type", "environment"]
        for param in kys:
            if param not in position:
                return {"err": f"Не корректно определён или не орпеделён параметр \'{data_mean[param]}\'"}

        if position["need_bellows"] is False:
            position["material_bellows"] = ""

        if position["valve_type"] == 'Н':
            position["spring_material"] = ""

        position["Pn"] = float(position["Pn"]) * 10.197162
        position["Pp"] = float(position["Pp"]) * 10.197162
        position["Pp_din"] = float(position["Pp"]) * 10.197162

        # Фланцы по ГОСТ 33259-2015
        if position["inlet_flange"] is not None:
            position["joining_type"] = position["joining_type"] + f" по ГОСТ 33259-2015 {position['inlet_flange']} / {position['outlet_flange']}"

        if position["valve_type"] == 'Н' or (
                position["valve_type"] == 'В' and position["open_close_type"] == "открытого типа") or (
                position["valve_type"] == 'В' and position["need_bellows"]):
            # Давление настройки без противодавления
            sheet[f"AL{i}"].value = position["Pn"]

            # Давление начала открытия без противодавления
            sheet[f"AM{i}"].value = position["Pno"]

            # Давление полного открытия без противодавления
            sheet[f"AN{i}"].value = position["Ppo"]

            # Давление настройки с противодавлением
            sheet[f"AH{i}"].value = position["Pn"]

            # Давление начала открытия с противодавлением
            sheet[f"AI{i}"].value = position["Pno"]

            # Давление полного открытия с противодавлением
            sheet[f"AJ{i}"].value = position["Ppo"]
        else:
            # Давление настройки без противодавления
            sheet[f"AL{i}"].value = position["Pn"]

            # Давление начала открытия без противодавления
            sheet[f"AM{i}"].value = position["Pno"]

            # Давление полного открытия без противодавления
            sheet[f"AN{i}"].value = position["Ppo"]

            # Давление настройки с противодавлением
            sheet[f"AH{i}"].value = position["Pn"] - position["Pp"]

            # Давление начала открытия с противодавлениемпротиводавлением
            sheet[f"AI{i}"].value = position["Pno"] - position["Pp"]

            # Давление полного открытия с противодавлением
            sheet[f"AJ{i}"].value = position["Ppo"] - position["Pp"]

        # номерация
        sheet[f"A{i}"].value = int(sheet[f"A3"].value) + i - 3

        # изготовитель
        sheet[f"BB{i}"].value = sheet[f"BB3"].value

        # Назначение
        sheet[f"D{i}"].value = "Общепромышленное"

        # Номер документа
        # print(i)
        sheet[f"E{i}"].value = "ТУ 3742-003-38877941-2012Б" if position[
                                                                   "valve_type"] == 'В' else "ТУ 3742-013-38877941-2016"

        type_name = "Пружинный" if position["valve_type"] == 'В' else "Пилотный"
        do = "прямого действия, " if position["valve_type"] == 'В' else ""
        # Тип клапана
        sheet[f"K{i}"].value = f"{type_name} предохранительный, сбросной, угловой, {do}{position['open_close_type']}"

        # Узел подрыва недоступен для заказа
        sheet[f"M{i}"].value = "Нет"

        # Коэффициент расхода, α
        alp = position["environment"] == "Газ"
        sheet[f"S{i}"].value = 0.8 if alp else 0.6

        '''
        #Диапазон настройки, кгс/см²
        if int(position["PN"]) == 16:
            sheet[f"AG{i}"].value = "0,5...16"
        elif position["PN"] == 25:
            sheet[f"AG{i}"].value = "6...25"
        elif position["PN"] == 40:
            sheet[f"AG{i}"].value = "8...40"
        elif position["PN"] == 63:
            sheet[f"AG{i}"].value = "16...63"
        elif position["PN"] == 100:
            sheet[f"AG{i}"].value = "40...100"
        elif position["PN"] == 160:
            sheet[f"AG{i}"].value = "40...160"
        '''

        # Гарантийный срок службы, мес.
        sheet[f"AY{i}"].value = 12

        # T окр среды
        sheet[f"J{i}"].value = f"{position['T_min']} ... {position['T_max']}"

        sheet[f"BB{i}"].value = "АО \"НПО Регулятор\""

        # заполнение по словарю
        for key in data_keys.keys():
            if type(position[data_keys[key]]) == type(True):
                if position[data_keys[key]] == True:
                    sheet[f"{key}{i}"].value = "Да"
                elif position[data_keys[key]] == False:
                    sheet[f"{key}{i}"].value = "Нет"
            elif position[data_keys[key]] == "" or position[data_keys[key]] == None:
                sheet[f"{key}{i}"].value = "Нет"
            else:
                sheet[f"{key}{i}"].value = position[data_keys[key]]

    # Создать экземпляр файла
    WB.save("./TKPexample.xlsx")

    return True


def make_OL(data):
    wb = load_workbook("./src/ОЛ.xlsx")
    sheet = wb['Table 1']

    # Трассировка данных
    params = {
        3: "OL_num",
        5: "quantity",
        6: "valve_type",  # переделать в слова
        7: "environment",
        8: "name",
        9: "T",
        10: "abrasive_particles",
        11: "density",
        12: "molecular_weight",  # если есть
        13: "adiabatic_index",  # если есть
        14: "viscosity",  # если жидкость
        15: "Pn",
        16: "Pno",
        17: "Ppo",
        18: "Pp",
        19: "Gab",
        20: "pre_Kc",
        21: "DN_s",
        22: "joining_type",
        23: "inlet_flange",
        24: "outlet_flange",
        25: "need_bellows",
        26: "force_open",
        27: "DN",
        28: "DN2",
        29: "PN",
        30: "PN2",
        31: "climate",
        32: "Tokr",
        33: "tightness",
        34: "material",
        35: "need_ZIP",
        36: "needKOF",
        37: "trials",
        38: "color",
        39: "packaging",
        40: "acceptance",
        41: "additionally"
    }

    '''Форматирование данных'''

    valve_type = "Пружинный" if data["valve_type"] == 'В' else "Пилотный"
    data["valve_type"] = valve_type

    # с противодавлением и перевести в МПа => /10
    Pn = float(data["Pn"]) * 0.1 - float(data['Pp']) * 0.1

    data['Pn'] = Pn

    #Фланцы по ГОСТ 33259-2015
    if data["inlet_flange"] is not None:
        data["inlet_flange"] = data["inlet_flange"] + " по ГОСТ 33259-2015"
        data["outlet_flange"] = data["outlet_flange"] + " по ГОСТ 33259-2015"

    # Давление начала открытия
    if Pn <= 0.3:
        Pno = Pn + 0.02
        data['Pno'] = Pno
    elif (Pn > 0.3) and (Pn <= 6):
        Pno = 1.07 * Pn
        data['Pno'] = Pno
    elif Pn > 6:
        Pno = 1.05 * Pn
        data['Pno'] = Pno

    # Давление полного открытия
    if Pn <= 0.3:
        Ppo = Pn + 0.05
        data['Ppo'] = Ppo
    elif (Pn > 0.3) and (Pn <= 6):
        Ppo = 1.15 * Pn
        data['Ppo'] = Ppo
    elif Pn > 6:
        Ppo = 1.1 * Pn
        data['Ppo'] = Ppo

    # Перевести в МПа => /10
    Pp = data["Pp"] * 0.1
    data['Pp'] = Pp

    Tokr = f"{data['T_min']} ... {data['T_max']}"
    data['Tokr'] = Tokr

    '''Автозаполнение'''
    for i in params.keys():
        # Проверка
        if params[i] not in data:
            return {"err": f"Не корректно определён или не орпеделён параметр \'{data_mean[params[i]]}\'"}
        else:
            if type(data[params[i]]) == type(True):
                if data[params[i]] == True:
                    sheet[f"C{i}"] = "Да"
                elif data[params[i]] == False:
                    sheet[f"C{i}"] = "Нет"
            elif data[params[i]] == "" or data[params[i]] == None:
                sheet[f"C{i}"] = "Нет"
            else:
                sheet[f"C{i}"] = data[params[i]]

    # Создать экземпляр файла
    wb.save("./OLexample.xlsx")

    return True
