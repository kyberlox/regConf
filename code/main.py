from fastapi import FastAPI, Body, Cookie, Header, Response, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware

from src.Raschet import Raschet, mixture, mark_params, get_tightness, make_XL, make_OL
from src.User import User

from openpyxl import load_workbook

from sqlalchemy.orm import sessionmaker
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy import create_engine, Column, Integer, Text, Float

import psycopg2

import json

import os
from dotenv import load_dotenv

import redis

import requests

#from typing import Optional



load_dotenv()

user = os.getenv('user')
pswd = os.getenv('pswd')
port = os.getenv('PORT')



def DB_exec(command):
    conn = psycopg2.connect(dbname="pdb", host="postgres", user="kyberlox", password="4179", port="5432")
    cursor = conn.cursor()
    cursor.execute(command)
    conn.commit()
    conn.close()

def DB_fetchOne(command):
    conn = psycopg2.connect(dbname="pdb", host="postgres", user="kyberlox", password="4179", port="5432")
    cursor = conn.cursor()
    cursor.execute(command)
    answer = cursor.fetchone()[0]
    conn.close()
    return answer

def DB_fetchAll(command):
    conn = psycopg2.connect(dbname="pdb", host="postgres", user="kyberlox", password="4179", port="5432")
    cursor = conn.cursor()
    cursor.execute(command)
    answer = cursor.fetchall()
    conn.close()
    return answer


engine = create_engine(f'postgresql+psycopg2://{user}:{pswd}@postgres/pdb')

class Base(DeclarativeBase): pass

class Table(Base):
    __tablename__ = 'environment_table'
    id = Column(Integer, primary_key=True)
    name = Column(Text)
    environment = Column(Text)
    molecular_weight = Column(Float, nullable=True)
    density = Column(Float, nullable=True)
    material = Column(Text, nullable=True)
    viscosity = Column(Float, nullable=True)
    isobaric_capacity = Column(Float, nullable=True)
    molar_mass = Column(Float, nullable=True)
    isochoric_capacity = Column(Float, nullable=True)
    adiabatic_index = Column(Float, nullable=True)
    compressibility_factor = Column(Float, nullable=True)

class Params(Base):
    __tablename__ = 'parametrs_table'
    id = Column(Integer, primary_key=True)
    DNS = Column(Float, nullable=True)
    Pnd = Column(Text, nullable=True)
    DN = Column(Float, nullable=True)
    PN = Column(Float, nullable=True)
    spring_material = Column(Text, nullable=True)
    spring_number = Column(Text, nullable=True)
    valve_type = Column(Text, nullable=True)

class Table2(Base):
    __tablename__ = 'table2'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)

class Table10(Base):
    __tablename__ = 'table10'
    id = Column(Integer, primary_key=True)
    T = Column(Float, nullable=True)
    Pn = Column(Float, nullable=True)
    P = Column(Float, nullable=True)

class pakingParams(Base):
    __tablename__ = 'paking_params'
    id = Column(Integer, primary_key=True)
    mark = Column(Text, nullable=True)
    DN = Column(Float, nullable=True)
    PN = Column(Float, nullable=True)
    M = Column(Float, nullable=True)
    S = Column(Float, nullable=True)

Base.metadata.create_all(bind=engine)
SessionLocal = sessionmaker(autoflush=True, bind=engine)
db = SessionLocal()



r = redis.Redis(host='redis', port=6379, username=user, password=pswd, db=0)



app = FastAPI()



origins = [
    "http://localhost:8000",
    "http://localhost:5173",
    "http://regconf.emk.ru",
    "https://localhost:8000",
    "https://localhost:5173",
    "https://regconf.emk.ru",
    "https://portal.emk.ru",
    "http://10.34.172.121:5173",
    "http://213.87.71.131",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=["GET", "POST", "DELETE", "PUT", "OPTIONS", "PATH"],
    allow_headers=["*"]
    #allow_headers=["Content-Type", "Accept", "Authorization", "Location", "Allow", "Content-Disposition", "Sec-Fetch-Dest", "Access-Control-Allow-Credentials"],
)

#app.mount("/", StaticFiles(directory="../front/dist", html=True), name="static")

#Функция для получения данных о пользователе по session_id
def check_session_id(token):
    url = "http://intranet.emk.org.ru/api/auth_router/check"
    # url = "https://intranet.emk.ru/api/auth_router/check"
    cookies = { 'session_id': token}
    res = requests.get(url, cookies=cookies)
    return json.loads(res.text)

#миграция и таблицы эксель
@app.get("/api/migration", tags=["Подбор"])
def migration():

    #прочитать из таблицы
    wb = load_workbook("./files/table.xlsx")
    sheet = wb['table']
    column_names = [
        "name",
        "environment",
        "molecular_weight",
        "density",
        "material",
        "viscosity",
        "isobaric_capacity",
        "molar_mass",
        "isochoric_capacity",
        "adiabatic_index",
        "compressibility_factor"
    ]
    cell_names = ['A', 'B', 'C', 'D', 'E',  'F', 'G',  'H', 'I',  'J', 'K']
    
    result = []
    for i in range(3, sheet.max_row+1):
        row = dict()
        for j in range(len(column_names)):
            row[f"{column_names[j]}"] = sheet[f"{cell_names[j]}{i}"].value
        result.append(row)
    
    #записать в БД
    for env in result:
        #построчная запись после проверки
        line = Table(name = env['name'], environment = env['environment'], molecular_weight = env['molecular_weight'], density = env['density'], material = env['material'], viscosity = env['viscosity'], isobaric_capacity = env['isobaric_capacity'], molar_mass = env['molar_mass'], isochoric_capacity = env['isochoric_capacity'], adiabatic_index = env['adiabatic_index'], compressibility_factor = env['compressibility_factor'])
        #lines = select(Table).where(Table.name == env['name'], Table.environment == env['environment'])
        lines = db.query(Table).filter(Table.name == env['name'], Table.environment == env['environment']).first()
        
        if lines == None:
            db.add(line)
            db.commit()
            env['id'] = line.id
        #перезаписать в БД
        else:
            line = db.query(Table).filter(Table.name == env['name'] and Table.environment == env['environment']).first()
            params = [line.environment, line.molecular_weight, line.density, line.material, line.viscosity, line.isobaric_capacity, line.molar_mass, line.isochoric_capacity, line.adiabatic_index, line.compressibility_factor]
            keys = ["environment", "molecular_weight", "density", "material", "viscosity", "isobaric_capacity", "molar_mass", "isochoric_capacity", "adiabatic_index", "compressibility_factor"]
            need = False
            for i in range(len(keys)):
                if params[i] != env[keys[i]]:
                    params[i] = env[keys[i]]
                    need = True
            if need:
                db.commit()
    
    #миграция параметров DN и PN
    WB = load_workbook("./files/PNtoDN.xlsx")
    sheet = WB['full']

    par_result = {"added" : [], "exists" : []}

    for i in range(2, sheet.max_row+1):
        DNS = float(sheet[f"A{i}"].value)
        P1_max = str(sheet[f"D{i}"].value)
        DN = float(sheet[f"C{i}"].value)
        PN = float(sheet[f"B{i}"].value)
        spring_material = str(sheet[f"F{i}"].value)
        spring_number = str(sheet[f"E{i}"].value)
        valve_type = str(sheet[f"H{i}"].value)

        #экземпляр таблицы параметров
        example = Params(DNS = DNS, Pnd = P1_max, DN = DN, PN = PN, spring_material = spring_material, spring_number = spring_number, valve_type = valve_type)
        
        #есть ли такая запись?
        request = db.query(Params).filter(Params.DNS == DNS, Params.Pnd == P1_max, Params.DN == DN, Params.PN == PN, Params.spring_material == spring_material, Params.spring_number == spring_number, Params.valve_type == valve_type).first()
        #print(request)

        #если нет - добавить
        if request == None:
            db.add(example)
            db.commit()

            curr = {
                "№" : i,
                "ID" : example.id,  
                "DNS" : DNS, 
                "Pnd" : P1_max, 
                "DN" : DN, 
                "PN" : PN, 
                "spring_material" : spring_material,
                "spring_number" : spring_number,
                "valve_type" : valve_type
            }
            par_result["added"].append(curr)

        #если есть - пропустить
        else:
            curr = {
                "ID" : request.id,  
                "DNS" : request.DNS, 
                "Pnd" : request.Pnd, 
                "DN" : request.DN, 
                "PN" : request.PN, 
                "spring_material" : request.spring_material,
                "spring_number" : spring_number,
                "valve_type" : valve_type
            }
            par_result["exists"].append(curr)

    wb1 = load_workbook("./files/Table2.xlsx")
    sheet = wb1['Лист1']

    t2_result = {"added" : [], "exists" : []}
    for i in range(2, 49):
        T = float(sheet[f"A{i}"].value)
        Pn = float(sheet[f"B{i}"].value)
        P = float(sheet[f"C{i}"].value)

        #print(i, T, Pn, P)

        example = Table2(T=T, Pn=Pn, P=P)
        request = db.query(Table2).filter(Table2.T == T, Table2.Pn == Pn, Table2.P == P).first()

        #если нет - добавить
        if request == None:
            db.add(example)
            db.commit()

            curr = {
                "№" : i,
                "ID" : example.id,  
                "T" : T, 
                "Pn" : Pn, 
                "P" : P
            }
            t2_result["added"].append(curr)
        #если есть - пропустить
        else:
            curr = {
                "ID" : request.id,  
                "T" : request.T, 
                "Pn" : request.Pn, 
                "P" : request.P
            }
            t2_result["exists"].append(curr)
        
    result.append(par_result)

    wb2 = load_workbook("./files/Table10.xlsx")
    sheet = wb2['Лист1']

    t10_result = {"added" : [], "exists" : []}

    for i in range(2, 72):
        T = float(sheet[f"A{i}"].value)
        Pn = float(sheet[f"B{i}"].value)
        P = float(sheet[f"C{i}"].value)

        example = Table10(T=T, Pn=Pn, P=P)
        request = db.query(Table10).filter(Table10.T == T, Table10.Pn == Table10.Pn, Table10.P == P).first()
        
        #если нет - добавить
        if request == None:
            db.add(example)
            db.commit()

            curr = {
                "№" : i,
                "ID" : example.id,  
                "T" : T, 
                "Pn" : Pn, 
                "P" : P
            }
            t10_result["added"].append(curr)
        #если есть - пропустить
        else:
            curr = {
                "ID" : request.id,  
                "T" : request.T, 
                "Pn" : request.Pn, 
                "P" : request.P
            }
            t10_result["exists"].append(curr)

    result.append(t10_result)

    wb3 = load_workbook("./files/paking_params.xlsx")
    sheet = wb3["Лист1"]

    pak_res = {"added" : [], "exists" : []}
    for i in range(2, sheet.max_row):
        mark = str(sheet[f"A{i}"].value)
        DN = float(sheet[f"B{i}"].value)
        PN = float(sheet[f"C{i}"].value)
        M = sheet[f"D{i}"].value
        S = sheet[f"E{i}"].value
        if type(M) == type("") or type(M) == type(1.0):
            M = float(sheet[f"D{i}"].value)
        else:
            M = None
        if type(S) == type("") or type(S) == type(1.0):
            S = float(sheet[f"E{i}"].value)
        else:
            S = None



        example = pakingParams(mark=mark, DN=DN, PN=PN, M=M, S=S)
        request = db.query(pakingParams).filter(pakingParams.mark == mark, pakingParams.DN == DN, pakingParams.PN == PN).first()

        #если нет - добавить
        if request == None:
            db.add(example)
            db.commit()

            curr = {
                "№" : i,
                "ID" : example.id,  
                "mark" : mark, 
                "DN" : DN, 
                "PN" : PN,
                "M" : M,
                "S" : S
            }
            pak_res["added"].append(curr)
        #если есть - пропустить
        else:
            curr = {
                "№" : i,
                "ID" : example.id,  
                "mark" : mark, 
                "DN" : DN, 
                "PN" : PN,
                "M" : M,
                "S" : S
            }
            pak_res["exists"].append(curr)



    return {"environmentTable" : result, "valveParametrsTable" : par_result, "Table2" : t2_result, "Table10" : t10_result, "PackingParams" : pak_res}

#подбор сред
@app.get("/api/get_table", tags=["Подбор"])
async def get_table():
    #вывод словаря
    environments = []
    lines = db.query(Table).all()
    for line in lines:
        environment = {
            "id" : line.id,
            "name" : line.name,
            "environment" : line.environment,
            "molecular_weight" : line.molecular_weight,
            "density" : line.density,
            "material" : line.material,
            "viscosity" : line.viscosity,
            "isobaric_capacity" : line.isobaric_capacity,
            "molar_mass" : line.molar_mass,
            "isochoric_capacity" : line.isochoric_capacity,
            "adiabatic_index" : line.adiabatic_index,
            "compressibility_factor" : line.compressibility_factor
        }
            
        environments.append(environment)
    
    return environments

#подбор смесей
@app.post("/api/get_compound", tags=["Подбор"])
async def get_compound(data=Body()):
    environments = []
    lines = db.query(Table).all()

    for env in data:
        if "id" in env and "r" in env:
            for line in lines:
                ID = env["id"]
                r = env["r"]
                if line.id == ID:
                    environment = {
                        "id": line.id,
                        "name": line.name,
                        "environment": line.environment,
                        "molecular_weight": line.molecular_weight,
                        "density": line.density,
                        "material": line.material,
                        "viscosity": line.viscosity,
                        "isobaric_capacity": line.isobaric_capacity,
                        "molar_mass": line.molar_mass,
                        "isochoric_capacity": line.isochoric_capacity,
                        "adiabatic_index": line.adiabatic_index,
                        "compressibility_factor": line.compressibility_factor,
                        "r": r
                    }

                    print(line.density, r)

                    environments.append(environment)
        elif "climate" in env:
            climate = env["climate"]
        elif "T" in env:
            T = env["T"]

    return mixture(environments, climate, T)

#получение осатльных параметров
@app.post("/api/get_pressure", tags=["Подбор"])
def get_pressure(data = Body()):
    for key in data.keys():
        if ((data[key] == "") or (data[key] == None)):
            return {"error" : "incorrect value", "key" : key, "value" : data[key]}
        else:
            #расчет
            return Raschet(data)

#подбор оборудования
@app.post("/api/get_mark_params", tags=["Подбор"])
async def get_mark_params(data = Body()):
    return mark_params(data)

@app.post("/api/get_tightness", tags=["Подбор"])
async def web_get_tightness(data = Body()):
    return get_tightness(data)



#авторизазия => генерация токена, начало сессии
# @app.post("/api/auth", tags=["Активность пользователей"])
# async def login(jsn = Body()):
#     print(jsn)
#     uuid = jsn["uuid"]
#     fio = f"{jsn['fio'][1]} {jsn['fio'][0]} {jsn['fio'][2]}"
#     dep = ""
#     for dp in jsn["department"]:
#         dep += dp
#     #запрос на БД
#     usr = User(uuid=uuid, fio=fio, department=dep)
#     tkn = usr.authenticate()

#     return {"token" : tkn}

@app.post("/api/auth", tags=["Активность пользователей"])
async def login(token: str):
    # token = request.cookies["session_id"]
    print(token, 'token')
    token_data = check_session_id(token)
    print(token_data, 'че приходит')
    user_info = token_data['user']
    uuid = user_info['uuid']
    fio = user_info['full_name']
    dep = ""
    for dp in user_info["department"]:
        dep += dp
    #запрос на БД
    usr = User(uuid=uuid, fio=fio, department=dep)
    tkn = usr.authenticate(sess_token=token)

    return {"token" : token}




#проверка авторизациии
@app.post("/api/check", tags=["Активность пользователей"])
async def check_valid(token: str = Header(None)):
    print(token)

    if token[:3] == "ip:":
        ip = token[3:]
        usr = User(ip=ip)
        usr_token = usr.authenticate()

        result = usr.check()
        if result is None:
            return {"error" : "invalid token"}
        else:
            content = {"token_valid": usr.check()}
            return JSONResponse(content=content, headers={"token": usr_token})

    else:
        # usr = User(token=token)

        # result = usr.check()
        result = check_session_id(token)
        if result is None:
            return {"error" : "invalid token"}
        else:
            return {"token_valid" : result}




#записать json в Redis
@app.post("/api/set_data", tags=["Активность пользователей"])
def get_data(data = Body(), token = Header(None)):
    usr = User(token=token, jsn=data)
    usr.set_dt()

    return usr.get_dt()

#получить json из Redis
@app.get("/api/get_data", tags=["Активность пользователей"])
async def get_data(token = Header(default=None)):
    usr = User(token=token)
    return usr.get_dt()

#прекратить сессию -> выйти из Redis
@app.get("/api/outh", tags=["Активность пользователей"])
async def outh_user(token = Header(default=None)):
    usr = User(token=token)
    usr.outh()
    return {'status' : 'ready'}



@app.post("/api/history", tags=["Активность пользователей"])
async def get_history(token = Header(None)):
    print(token, 'token из Header')
    usr = User(token=token)
    if token[:3] != "ip:":
        return usr.history()
    if usr.check():
        return usr.history()
    else:
        return {"error" : "invalid token"}

@app.delete("/api/delete_tkp/{tkp_id}", tags=["Активность пользователей"])
async def delete_tkp_id(tkp_id, token = Header(None)):
    usr = User(token=token)
    if token[:3] != "ip:":
        return usr.deleteConfiguration(tkp_id)
    if usr.check():
        return usr.deleteConfiguration(tkp_id)

@app.get("/api/upload_tkp/{tkp_id}", tags=["Активность пользователей"])
async def upload_tkp_id(tkp_id, token = Header(None)):
    usr = User(token=token)
    if token[:3] != "ip:":
        return usr.uploadConfiguration(tkp_id)
    if usr.check():
        return usr.uploadConfiguration(tkp_id)

@app.post("/api/add_position_tkp", tags=["Активность пользователей"])
async def add_position_tkp_id(tkp_position = Body(), token = Header(None)):
    usr = User(token=token)
    if token[:3] != "ip:":
        return usr.addPosition(tkp_position)
    if usr.check():
        return usr.addPosition(tkp_position)

@app.delete("/api/delete_position_tkp/{tkp_id}/{position_id}", tags=["Активность пользователей"])
async def delete_position_tkp_id(tkp_id : int, position_id : int, token = Header(None)):
    usr = User(token=token)
    if token[:3] != "ip:":
        return usr.deletePosition(tkp_id, position_id)
    if usr.check():
        return usr.deletePosition(tkp_id, position_id)



#генерация документации
@app.post("/api/generate", tags=["Генерация документации"]) #проверка сессии
def generate(data = Body(), name = Header(default=None) , token = Header(default=None)):
    usr = User(token=token)
    if token[:3] != "ip:":
        jsn = usr.create_TKP(name)
    else:
        if usr.check():
            # получить название и сохранить в БД
            #получить json для генерации из Redis
            jsn = usr.create_TKP(name)
        else:
            jsn = data

    #генерация файла
    res = make_XL(jsn)

    #выдать файл
    if res == True:
        return FileResponse("./TKPexample.xlsx", filename=f"{name} ТКП ПК.xlsx", media_type="application/xlsx")
    else:
        return res


@app.post("/api/makeOL", tags=["Генерация документации"])
def generate_OL(data = Body(), token: str = Header(None)):
    #запись в БД
    if token[:3] == "ip:":
        ip = token[3:]
        usr = User(ip=ip)
        token = usr.authenticate()

    usr = User(token=token, jsn=data)

    if usr.create_OL():

        #сохранить json
        #f = open(f"./data/OL.json", 'w')
        #json.dump(data, f)
        #f.close()

        #генерация файла
        res = make_OL(data)
        #return res
        #выдать файл
        if res:
            return FileResponse(f'./OLexample.xlsx', filename=f'ОЛ ПК.xlsx', media_type='application/xlsx')
        else:
            return res
